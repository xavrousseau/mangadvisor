from __future__ import annotations

import csv
import io
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any


API_BASE_URL = os.getenv("MANGADVISOR_API_BASE_URL", "http://localhost:8000")
ROOT_DIR = Path(__file__).resolve().parents[3]
REPORT_DIR = ROOT_DIR / "docs" / "reports"
REPORT_PATH = REPORT_DIR / "library_import_tests_v0_8_10.md"


IMPORT_ROWS: list[dict[str, Any]] = [
    {
        "title": "Naruto",
        "library_status": "READ",
        "user_score": "8.5",
        "is_favorite": "true",
        "owned_volumes": "72",
        "read_volumes": "72",
        "notes": "Très bon shonen",
    },
    {
        "title": "Bleach",
        "library_status": "READ",
        "user_score": "8",
        "is_favorite": "false",
        "owned_volumes": "74",
        "read_volumes": "74",
        "notes": "",
    },
    {
        "title": "One Piece",
        "library_status": "WANT_TO_READ",
        "user_score": "",
        "is_favorite": "false",
        "owned_volumes": "",
        "read_volumes": "",
        "notes": "",
    },
    {
        "title": "Fairy Tail",
        "library_status": "NOT_INTERESTED",
        "user_score": "",
        "is_favorite": "false",
        "owned_volumes": "",
        "read_volumes": "",
        "notes": "",
    },
    {
        "title": "Titre Qui Nexiste Pas 999",
        "library_status": "READ",
        "user_score": "7",
        "is_favorite": "false",
        "owned_volumes": "",
        "read_volumes": "",
        "notes": "Doit ressortir en non trouvé",
    },
]


FIELDNAMES = [
    "title",
    "library_status",
    "user_score",
    "is_favorite",
    "owned_volumes",
    "read_volumes",
    "notes",
]


def api_request(
    path: str,
    method: str = "GET",
    payload: dict[str, Any] | None = None,
    params: dict[str, Any] | None = None,
) -> dict[str, Any]:
    url = f"{API_BASE_URL}{path}"

    if params:
        url = f"{url}?{urllib.parse.urlencode(params)}"

    data = None
    headers = {}

    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"

    request = urllib.request.Request(
        url=url,
        data=data,
        headers=headers,
        method=method,
    )

    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            body = response.read().decode("utf-8")
            return json.loads(body) if body else {}

    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Erreur HTTP {exc.code} sur {method} {path}: {body}") from exc

    except urllib.error.URLError as exc:
        raise RuntimeError(
            f"Impossible d'appeler l'API {API_BASE_URL}. Vérifie que mangadvisor_api est lancé."
        ) from exc


def api_get(path: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
    return api_request(path=path, method="GET", params=params)


def api_post(path: str, payload: dict[str, Any]) -> dict[str, Any]:
    return api_request(path=path, method="POST", payload=payload)


def api_delete(path: str) -> dict[str, Any]:
    return api_request(path=path, method="DELETE")


def build_csv_content(rows: list[dict[str, Any]]) -> str:
    output = io.StringIO()

    writer = csv.DictWriter(output, fieldnames=FIELDNAMES)
    writer.writeheader()
    writer.writerows(rows)

    return output.getvalue()


def build_csv_content_from_excel_if_possible(rows: list[dict[str, Any]]) -> tuple[str | None, str | None]:
    """
    Teste la compatibilité Excel sans passer par Streamlit :
    - crée un fichier Excel en mémoire ;
    - le relit ;
    - reconvertit les lignes en CSV ;
    - réutilise l'endpoint /library/import/csv.

    Si openpyxl n'est pas installé localement, le test Excel est ignoré.
    """
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        return None, "openpyxl non installé localement : test Excel ignoré."

    excel_buffer = io.BytesIO()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "bibliotheque"

    worksheet.append(FIELDNAMES)

    for row in rows:
        worksheet.append([row.get(field, "") for field in FIELDNAMES])

    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    loaded_workbook = load_workbook(excel_buffer)
    loaded_sheet = loaded_workbook.active

    loaded_rows = list(loaded_sheet.iter_rows(values_only=True))

    if not loaded_rows:
        raise RuntimeError("Le fichier Excel de test est vide.")

    headers = [str(value) for value in loaded_rows[0]]
    normalized_rows: list[dict[str, Any]] = []

    for values in loaded_rows[1:]:
        normalized_rows.append(
            {
                headers[index]: "" if value is None else value
                for index, value in enumerate(values)
            }
        )

    return build_csv_content(normalized_rows), None


def clear_library() -> None:
    data = api_get("/library", params={"limit": 500})
    items = data.get("items", [])

    for item in items:
        manga_id = item.get("manga_id")

        if manga_id:
            api_delete(f"/library/items/{manga_id}")


def import_csv_content(csv_content: str, dry_run: bool) -> dict[str, Any]:
    return api_post(
        "/library/import/csv",
        payload={
            "csv_content": csv_content,
            "delimiter": ",",
            "default_status": "WANT_TO_READ",
            "dry_run": dry_run,
        },
    )


def get_library_items() -> list[dict[str, Any]]:
    data = api_get("/library", params={"limit": 500})
    return data.get("items", [])


def get_library_profile() -> dict[str, Any]:
    return api_get("/library/profile")


def test_import_case(
    name: str,
    csv_content: str,
    dry_run: bool,
) -> dict[str, Any]:
    clear_library()

    result = import_csv_content(
        csv_content=csv_content,
        dry_run=dry_run,
    )

    library_items = get_library_items()
    profile = get_library_profile()

    return {
        "name": name,
        "dry_run": dry_run,
        "import_result": result,
        "library_items": library_items,
        "profile": profile,
    }


def md_cell(value: Any) -> str:
    if value is None:
        return ""

    return str(value).replace("|", "\\|").replace("\n", " ")


def test_status(case: dict[str, Any]) -> str:
    result = case["import_result"]
    library_items = case["library_items"]
    dry_run = case["dry_run"]

    matched_count_ok = result.get("matched_count") == 4
    not_found_ok = result.get("not_found_count") == 1
    error_count_ok = result.get("error_count") == 0

    if dry_run:
        library_count_ok = len(library_items) == 0
    else:
        library_count_ok = len(library_items) == 4

    if matched_count_ok and not_found_ok and error_count_ok and library_count_ok:
        return "OK"

    return "KO"


def main() -> int:
    if "--reset-library" not in sys.argv:
        print("ERREUR : ce script vide la bibliothèque locale avant les tests.")
        print("Relance avec :")
        print("  python pipelines\\engine\\scripts\\run_library_import_tests.py --reset-library")
        return 1

    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    print("==================================================")
    print("Mangadvisor - Tests import bibliotheque")
    print("==================================================")
    print(f"API : {API_BASE_URL}")
    print()

    csv_content = build_csv_content(IMPORT_ROWS)

    cases: list[dict[str, Any]] = []

    cases.append(
        test_import_case(
            name="CSV simulation",
            csv_content=csv_content,
            dry_run=True,
        )
    )

    cases.append(
        test_import_case(
            name="CSV import réel",
            csv_content=csv_content,
            dry_run=False,
        )
    )

    excel_csv_content, excel_skip_reason = build_csv_content_from_excel_if_possible(IMPORT_ROWS)

    if excel_csv_content:
        cases.append(
            test_import_case(
                name="Excel converti en CSV - import réel",
                csv_content=excel_csv_content,
                dry_run=False,
            )
        )

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines: list[str] = [
        "# Mangadvisor — Tests import bibliothèque V0.8.10",
        "",
        f"Date d’exécution : `{now}`",
        "",
        f"API testée : `{API_BASE_URL}`",
        "",
        "## Synthèse",
        "",
        "| Test | Dry run | Matched | Non trouvés | Erreurs | Taille bibliothèque après test | Résultat |",
        "|---|---|---:|---:|---:|---:|---|",
    ]

    for case in cases:
        result = case["import_result"]
        library_items = case["library_items"]

        lines.append(
            "| "
            f"{md_cell(case['name'])} | "
            f"{'Oui' if case['dry_run'] else 'Non'} | "
            f"{result.get('matched_count', 0)} | "
            f"{result.get('not_found_count', 0)} | "
            f"{result.get('error_count', 0)} | "
            f"{len(library_items)} | "
            f"{test_status(case)} |"
        )

    if excel_skip_reason:
        lines.extend(
            [
                "",
                "## Test Excel",
                "",
                f"Test Excel ignoré : `{excel_skip_reason}`",
            ]
        )

    lines.extend(
        [
            "",
            "## Détail",
            "",
        ]
    )

    for case in cases:
        result = case["import_result"]
        library_items = case["library_items"]
        profile = case["profile"]

        lines.extend(
            [
                f"## {case['name']}",
                "",
                f"**Résultat :** `{test_status(case)}`",
                "",
                "### Lignes trouvées",
                "",
                "| Ligne | Demandé | Trouvé | Statut | Note | Favori |",
                "|---:|---|---|---|---:|---|",
            ]
        )

        for item in result.get("imported_items", []):
            lines.append(
                "| "
                f"{item.get('line')} | "
                f"{md_cell(item.get('requested_title'))} | "
                f"{md_cell(item.get('matched_title'))} | "
                f"{md_cell(item.get('library_status'))} | "
                f"{md_cell(item.get('user_score'))} | "
                f"{'Oui' if item.get('is_favorite') else 'Non'} |"
            )

        lines.extend(
            [
                "",
                "### Titres non trouvés",
                "",
                "| Ligne | Titre | Statut demandé |",
                "|---:|---|---|",
            ]
        )

        not_found_items = result.get("not_found_items", [])

        if not_found_items:
            for item in not_found_items:
                lines.append(
                    "| "
                    f"{item.get('line')} | "
                    f"{md_cell(item.get('title'))} | "
                    f"{md_cell(item.get('library_status'))} |"
                )
        else:
            lines.append("|  | Aucun |  |")

        lines.extend(
            [
                "",
                "### Erreurs",
                "",
                "| Ligne | Erreur |",
                "|---:|---|",
            ]
        )

        error_items = result.get("error_items", [])

        if error_items:
            for item in error_items:
                lines.append(
                    "| "
                    f"{item.get('line')} | "
                    f"{md_cell(item.get('error'))} |"
                )
        else:
            lines.append("|  | Aucune |")

        lines.extend(
            [
                "",
                "### Bibliothèque après test",
                "",
                "| Titre | Statut | Note | Favori |",
                "|---|---|---:|---|",
            ]
        )

        if library_items:
            for item in library_items:
                lines.append(
                    "| "
                    f"{md_cell(item.get('title'))} | "
                    f"{md_cell(item.get('library_status'))} | "
                    f"{md_cell(item.get('user_score'))} | "
                    f"{'Oui' if item.get('is_favorite') else 'Non'} |"
                )
        else:
            lines.append("| Bibliothèque vide |  |  |  |")

        lines.extend(
            [
                "",
                "### Profil après test",
                "",
                f"**Statuts :** `{profile.get('status_counts', {})}`",
                "",
                f"**Sources positives :** `{profile.get('positive_source_count', 0)}`",
                "",
                f"**Signaux négatifs :** `{profile.get('negative_item_count', 0)}`",
                "",
            ]
        )

    lines.extend(
        [
            "## Note",
            "",
            "Ce test valide l'endpoint `/library/import/csv` et vérifie que l'import alimente correctement la bibliothèque et le profil de lecture.",
            "Le test Excel crée un fichier `.xlsx` en mémoire si `openpyxl` est installé localement, puis le reconvertit en CSV pour réutiliser la même logique métier.",
            "",
        ]
    )

    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")

    print("==================================================")
    print("Tests terminés")
    print(f"Rapport généré : {REPORT_PATH}")
    print("==================================================")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())